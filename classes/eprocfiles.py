import pyexcel as p
import pyexcel_xls
import pyexcel_xlsx
from openpyxl import load_workbook
import glob
import os
import re
from functions.aux_functions import connectDataBase

class EprocFiles():

    final_file_name = ""

    def __init__(self, estado):
        self.estado = estado
        self.processNumbers = []
        self.processParts = []
        self.activityDueDates = []
        self.activityDueDatesStringified = []
        self.dia = []
        self.mes = []
        self.ano = []


    def initFile(self):
        self.getFiles()
        self.formatExcelFile()
        self.formatProcessData()
        self.addProcessesToDataBase()


    def getFiles(self):
        self.file_name = glob.glob("*.xls")
        self.convertAndRenameFile()


    def convertAndRenameFile(self):
        file_name = self.file_name[0]
        self.final_file_name = file_name[:16] + self.estado + ".xls"
        os.rename(file_name, self.final_file_name)
        p.save_book_as(file_name = self.final_file_name, dest_file_name = f"intimacao{self.estado}.xlsx")
        os.remove(self.final_file_name)


    def formatExcelFile(self):
        self.wb = load_workbook(f"intimacao{self.estado}.xlsx")
        self.ws = self.wb.active

        for _ in range(2):
            self.ws.delete_rows(1)

        self.ws.delete_cols(2)

        for _ in range(5):
            self.ws.delete_cols(3)

        self.getExcelData()


    def getExcelData(self):

        for i in range(1, self.ws.max_row + 1):
            self.processNumbers.append(self.ws.cell(row = i, column = 1).value)

        for i in range(1, self.ws.max_row + 1):
            self.processParts.append(self.ws.cell(row = i, column = 2).value)

        for i in range(1, self.ws.max_row + 1):
            self.activityDueDates.append(self.ws.cell(row = i, column = 3).value)
    
        os.remove(f"intimacao{self.estado}.xlsx")

        
    def formatProcessData(self):
        self.stringifyDate()
        self.formatProcessesNames()


    def stringifyDate(self):

        for date in self.activityDueDates:
            self.activityDueDatesStringified.append(date.strftime("%x"))
        
        for date in self.activityDueDatesStringified:
            self.mes.append(date[:-6])
            self.dia.append(date[3:-3])
            self.ano.append(date[6:])


    def formatProcessesNames(self):
        for i, partes in enumerate(self.processParts):
            self.processParts[i] = self.regexString(partes)


    def regexString(self, string):
        string = re.sub("\n\n\s\(\d+\)", "", string)
        string = re.sub("Exequente \n", "", string)
        string = re.sub("Executado \n", "", string)
        string = re.sub("Autor \n", "", string)
        string = re.sub("Réu \n", "", string)
        string = re.sub("Notificante \n", "", string)
        string = re.sub("Notificado \n", "", string)
        string = re.sub("Requerente \n", "", string)
        string = re.sub("Requerido \n", "", string)
        string = re.sub("\s\n", "\n", string)

        return string


    def addProcessesToDataBase(self):
        con = connectDataBase()

        cur = con.cursor()

        cur.execute("CREATE TABLE IF NOT EXISTS agendadoreproc (num_processo CHAR(25), partes VARCHAR(500), data DATE, estado CHAR(8));")

        for i in range(len(self.processNumbers)):
            cur.execute(f"INSERT INTO agendadoreproc VALUES ('{self.processNumbers[i]}', '{self.processParts[i]}', '{self.activityDueDates[i]}', 'eproc-{self.estado.upper()}')")

        con.commit()
        cur.close()
        con.close()


    @classmethod
    def resetFilesAndTables(cls):

        local_path = os.getcwd()
        for item in os.listdir(local_path):
            if item.endswith(".xls") or item.endswith(".xlsx"):
                os.remove(os.path.join(local_path, item))

        con = connectDataBase()
        cur = con.cursor()
        cur.execute("DROP TABLE IF EXISTS agendadoreproc;")
        con.commit()
        cur.close()
        con.close()