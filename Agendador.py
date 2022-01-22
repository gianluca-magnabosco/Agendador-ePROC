import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter.font import BOLD
from tkinter.ttk import *
from tkinter.constants import CENTER, TOP, DISABLED
import openpyxl as xl
from openpyxl import Workbook, load_workbook
import pyexcel as p
import pyexcel_xls
import pyexcel_xlsx
from datetime import datetime
import pandas as pd
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import os
import time
from selenium.webdriver.common.by import By
import glob
import winshell
from win32com.client import Dispatch
import pickle
import datetime
from collections import namedtuple
from google_auth_oauthlib.flow import Flow, InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from google.auth.transport.requests import Request



# delete inconvenient, possible waste files
local_path = os.getcwd()
listardir = os.listdir(local_path)
for item in listardir:
    if item.endswith(".xls"):
        os.remove(os.path.join(local_path, item))
    if item.endswith(".xlsx"):
        os.remove(os.path.join(local_path,item))


# create desktop shortcut
desktop = winshell.desktop()
path = os.path.join(desktop, "Agendador ePROC.lnk")
target = os.path.join(local_path, "Agendador ePROC.exe")
wDir = local_path
icon = os.path.join(local_path, "Agendador ePROC.exe")

if not os.path.exists(path):
    shell = Dispatch('WScript.Shell')
    shortcut = shell.CreateShortCut(path)
    shortcut.Targetpath = target
    shortcut.WorkingDirectory = wDir
    shortcut.IconLocation = icon
    shortcut.save()



def create_service(client_secret_file, api_name, api_version, *scopes, prefix=''):
	CLIENT_SECRET_FILE = client_secret_file
	API_SERVICE_NAME = api_name
	API_VERSION = api_version
	SCOPES = [scope for scope in scopes[0]]

	cred = None
	working_dir = os.getcwd()
	token_dir = 'token files'
	pickle_file = f'token_{API_SERVICE_NAME}_{API_VERSION}{prefix}.pickle'

	# check if token dir exists first, if not, create the folder
	if not os.path.exists(os.path.join(working_dir, token_dir)):
		os.mkdir(os.path.join(working_dir, token_dir))
	if os.path.exists(os.path.join(working_dir, token_dir, pickle_file)):
		with open(os.path.join(working_dir, token_dir, pickle_file), 'rb') as token:
			cred = pickle.load(token)

	if not cred or not cred.valid:
		if cred and cred.expired and cred.refresh_token:
			cred.refresh(Request())
		else:
			flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
			cred = flow.run_local_server()
		with open(os.path.join(working_dir, token_dir, pickle_file), 'wb') as token:
			pickle.dump(cred, token)
	try:
		service = build(API_SERVICE_NAME, API_VERSION, credentials=cred)
		print(API_SERVICE_NAME, API_VERSION, 'service created successfully')
		return service
	except Exception as e:
		print(e)
		print(f'Failed to create service instance for {API_SERVICE_NAME}')
		os.remove(os.path.join(working_dir, token_dir, pickle_file))
		return None




CLIENT_SECRET_FILE = "client_secret.json"
API_NAME = 'calendar'
API_VERSION = 'v3'
SCOPES = ['https://www.googleapis.com/auth/calendar']



service = create_service(CLIENT_SECRET_FILE, API_NAME, API_VERSION, SCOPES)



# center window
def center_window(width=860,height=640):
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    root.geometry('%dx%d+%d+%d' % (width, height, x, y))


# close program confirmation
def on_closeroot():
    close = messagebox.askokcancel("Confirmação", "Tem certeza que deseja fechar o programa?")
    if close:
        root.destroy()

# main function
def runcode():
    # style
    status = tk.Label(text="Carregando... Aguarde",wraplength=200,font=('',9,'bold'),bg='white')
    status.pack()
    status.place(relx=0.5, rely=0.7, anchor=CENTER)
    status.update_idletasks()

    button1.configure(state = 'disabled')
    button1.update_idletasks()
    time.sleep(1)
    button1.configure(state = 'enabled')
    status.destroy()


    calendar_body = {
        'summary': 'Intimações ePROC',
        'timeZone': 'America/Sao_Paulo',
    }

    calendar_list = service.calendarList().list(pageToken=None,maxResults=10).execute()


    # delete already existing calendar
    for calendar_list_entry in calendar_list['items']:  
        if 'Intimações ePROC' in calendar_list_entry['summary']:
            id = calendar_list_entry['id'] 
            service.calendars().delete(calendarId=id).execute()


    # create new calendar
    service.calendars().insert(body=calendar_body).execute()


    # list calendars
    calendar_list = service.calendarList().list(pageToken=None,maxResults=6).execute()

    # get calendar id
    for calendar_list_entry in calendar_list['items']:
        if 'Intimações ePROC' in calendar_list_entry['summary']:
            id = calendar_list_entry['id'] 






    # ------------------------------------- SANTA CATARINA ----------------------------------

    # web browser download xls file
    # configuring web browser
    options = Options()
    login = "LOGIN"
    passwd = "PASSWORD"
    options.add_argument("start-maximized")
    options.add_argument("--headless")
    preferences = {"download.default_directory": local_path,
                    "directory_upgrade": True}
    options.add_experimental_option("prefs", preferences)                   

    # open web browser
    driverpath = "/chromedriver.exe"
    driver = webdriver.Chrome(executable_path = local_path + driverpath, options=options)
    driver.get("https://eproc1g.tjsc.jus.br/eproc/externo_controlador.php?acao=principal")
    time.sleep(1)

    # login
    loginxpath = driver.find_element(By.XPATH, '//*[@id="txtUsuario"]')
    loginxpath.send_keys(login)

    passwxpath = driver.find_element(By.XPATH, '//*[@id="pwdSenha"]')
    passwxpath.send_keys(passwd)

    loginbuttonxpath = driver.find_element(By.XPATH, '//*[@id="sbmEntrar"]')
    loginbuttonxpath.click()

    # get to file webpage
    scbuttonxpath = driver.find_element(By.XPATH, '//*[@id="tr0"]')
    scbuttonxpath.click()
    time.sleep(2)

    intimacoesxpath = driver.find_element(By.XPATH, '//*[@id="conteudoCitacoesIntimacoesSC"]/div[2]/table/tbody/tr[1]/td[2]/a')
    intimacoesxpath.click()
    time.sleep(1)

    newURl = driver.window_handles[1]
    driver.switch_to.window(newURl)

    # download file
    gerarplanilhaid = driver.find_element(By.ID, 'sbmPlanilha')
    gerarplanilhaid.click()
    time.sleep(2)
    driver.quit()

    # rename downloaded file
    filename = glob.glob('*.xls')
    filename = ''.join(filename)
    finalfilename = filename[:16] + '.xls'
    os.rename(filename, finalfilename)

    # convert to xlsx
    p.save_book_as(file_name= finalfilename,
                dest_file_name='intimacaosc.xlsx')

    os.remove(finalfilename)


    filename = ''
    finalfilename = ''


    # web browser download xls file PARANA
    # open web browser
    driver = webdriver.Chrome(executable_path = local_path + driverpath, options=options)
    driver.get("https://eproc1g.tjsc.jus.br/eproc/externo_controlador.php?acao=principal")
    time.sleep(1)

    # login
    loginxpath = driver.find_element(By.XPATH, '//*[@id="txtUsuario"]')
    loginxpath.send_keys(login)

    passwxpath = driver.find_element(By.XPATH, '//*[@id="pwdSenha"]')
    passwxpath.send_keys(passwd)

    loginbuttonxpath = driver.find_element(By.XPATH, '//*[@id="sbmEntrar"]')
    loginbuttonxpath.click()

    # get to file webpage
    prbuttonxpath = driver.find_element(By.XPATH, '//*[@id="tr1"]')
    prbuttonxpath.click()
    time.sleep(2)

    intimacoesxpath = driver.find_element(By.XPATH, '//*[@id="conteudoCitacoesIntimacoesSC"]/div[2]/table/tbody/tr[1]/td[2]/a')
    intimacoesxpath.click()
    time.sleep(1)

    newURl = driver.window_handles[1]
    driver.switch_to.window(newURl)

    # download file
    gerarplanilhaid = driver.find_element(By.ID, 'sbmPlanilha')
    gerarplanilhaid.click()
    time.sleep(2)
    driver.quit()

    # rename downloaded file
    filename = glob.glob('*.xls')
    filename = ''.join(filename)
    finalfilename = filename[:16] + '.xls'
    os.rename(filename, finalfilename)


    # convert to xlsx
    p.save_book_as(file_name= finalfilename,
                dest_file_name='intimacaopr.xlsx')

    
    os.remove(finalfilename)





    # insert events to google calendar function
    def insert_events(location, color):
        all_day_event_true_start = []
        all_day_event_true_end = []
        for i in range(0,max_rows):
            all_day_event_true_start.append("{}-{}-{}".format(ano[i],mes[i],dia[i]))
            all_day_event_true_end.append("{}-{}-{}".format(ano[i],mes[i],dia[i]))
            event_request_body = {
                'start':{
                    'date': all_day_event_true_start[i],
                    'timeZone': 'America/Sao_Paulo',
                },
                'end':{
                    'date': all_day_event_true_end[i],
                    'timeZone': 'America/Sao_Paulo',
                },
                'summary': subj[i],
                'description': subj[i],
                'location': location,
                'colorId': color,
                #'attendees':[
                #    {
                #        'email': 'rasderfarr_gaguigo1@hotmail.com',
                #        'optional': False,
                #        'responseStatus': 'accepted',
                #    }
                #],
                #'reminders': {
                #    'useDefault': False,
                #    'overrides':[
                #        {'method': 'email', 'minutes': 30},
                #    ]
                #}
            }
            service.events().insert(calendarId=id, body=event_request_body).execute()

    

    # load file, sheet
    wb = load_workbook('intimacaosc.xlsx')
    ws = wb.active


    # formatting excel sheet
    for i in range(0,2):
        ws.delete_rows(1)

    for i in range(0,2):
        ws.delete_cols(1)

    for i in range(0,5):
        ws.delete_cols(2)


    # row number variable
    max_rows = ws.max_row

    # append list variables
    subj = []
    date = []

    # copy subject
    for i in range(1,max_rows+1):
        subj.append(ws.cell(row = i, column = 1).value)

    # copy date
    for i in range(1,max_rows+1):
        date.append(ws.cell(row = i, column = 2).value)

    # stringify date
    i = 0
    while i < max_rows:
        date[i] = date[i].strftime("%x")
        i += 1



    # append lists
    mes = []
    dia = []
    ano = []


    # format month
    for i in range(0,max_rows):
        mes.append(date[i])
    mes = [x[:-6] for x in mes]


    # format day
    for i in range(0,max_rows):
        dia.append(date[i])
    dia = [x[3:-3] for x in dia]


    # format year
    for i in range(0,max_rows):
        ano.append(date[i])
    ano = ['20' + x[6:] for x in ano]


    # insert events
    sc = "ePROC-SC"
    insert_events(sc, 7)    


    # delete no longer needed files
    os.remove('intimacaosc.xlsx')


    # -------------------------------------- PARANA ----------------------------------------------------


    # load file, sheet
    wb = load_workbook('intimacaopr.xlsx')
    ws = wb.active


    # formatting excel chart
    for i in range(0,2):
        ws.delete_rows(1)


    for i in range(0,2):
        ws.delete_cols(1)

    for i in range(0,5):
        ws.delete_cols(2)


    # row number variable
    max_rows = ws.max_row

    # append list variables
    subj = []
    date = []


    # copy subject
    for i in range(1,max_rows+1):
        subj.append(ws.cell(row = i, column = 1).value)


    # copy date
    for i in range(1,max_rows+1):
        date.append(ws.cell(row = i, column = 2).value)

    # stringify date
    i = 0
    while i < max_rows:
        date[i] = date[i].strftime("%x")
        i += 1




    # append lists
    mes = []
    dia = []
    ano = []


    # format month
    for i in range(0,max_rows):
        mes.append(date[i])
    mes = [x[:-6] for x in mes]


    # format day
    for i in range(0,max_rows):
        dia.append(date[i])
    dia = [x[3:-3] for x in dia]


    # format year
    for i in range(0,max_rows):
        ano.append(date[i])
    ano = ['20' + x[6:] for x in ano]

    # insert events
    pr = "ePROC-PR"
    insert_events(pr, 11)


    # delete no longer needed files
    #os.remove('intimacaopr.xlsx')

    # ----------------------------------------------------------------------------------------------




    # dynamic GUI
    status1 = tk.Label(text="Concluído!",wraplength=200,font=('',9,'bold'),bg='white')
    status1.pack()
    status1.place(relx=0.5, rely=0.7, anchor=CENTER)


    # confirmation
    def on_closetop():
        status = tk.Label(text="As intimações foram importadas e se encontram no Google Agenda!",wraplength=200,font=('',9,'bold'),bg='white')
        status.pack()
        status.place(relx=0.5, rely=0.71, anchor=CENTER)
        status.update_idletasks()
        close = messagebox.showinfo("Informação", "As intimações foram adicionadas no Google Agenda com sucesso!")
        if close:
            close2 = messagebox.askyesno("Sair","Deseja sair do programa?")
            if close2:
                root.destroy()
            else:
                status.destroy()
                status.update_idletasks()

    # end program
    def endprogram():
        status1.destroy()
        on_closetop() 


    closeconfirmation = messagebox.showinfo("Sucesso!", "O programa foi executado com sucesso!")
    if closeconfirmation:
        endprogram()
    

    #
    #
    #
    # python -m PyInstaller --onedir --windowed --icon=icone.ico --name="Agendador ePROC" Agendador.py
    #
    #
    #

# static GUI
root = tk.Tk()
center_window(860, 640)
root.title("Agendador ePROC")
bg = PhotoImage(file = "background.png")
background_label = Label(root, image=bg)
background_label.place(relx=0.5,rely=0.5,anchor=CENTER)
root.protocol("WM_DELETE_WINDOW", on_closeroot)
iconFile = 'icone.ico'
root.iconbitmap(default=iconFile)

# menu instructions
info = tk.Label(text="Clique no botão abaixo para iniciar o programa e aguarde alguns instantes!",wraplength=200,font=('',13),bg='white')
info.pack()
info.place(relx=0.5, rely=0.4, anchor=CENTER)

# welcoming message
status = tk.Label(text="Seja bem-vindo!",wraplength=200,font=('',15,'bold'),bg='white')
status.pack()
status.place(relx=0.5, rely=0.52, anchor=CENTER)

# credits
feitopor = tk.Label(text="Programa criado por: Gianluca Notari Magnabosco da Silva",font=('',7),bg="white")
feitopor.pack()
feitopor.place(relx=0.85, rely=0.98, anchor=CENTER)

# run code button
st = Style()
st.configure('W.TButton', background='white', foreground='black', font=('Open Sans',11))
button1 = Button(root, style='W.TButton', text='Clique aqui para iniciar o programa',command=runcode,width=27.75)
button1.pack()
button1.place(relx=0.5, rely=0.65, anchor=CENTER)

# end of GUI
root.mainloop()